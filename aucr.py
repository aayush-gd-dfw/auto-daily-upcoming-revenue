

import os
import io
import re
import base64
from datetime import datetime, date, timezone
from typing import Optional, Dict, Any, List, Tuple

import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook


# -------------------- CONFIG --------------------
TAB_NAME = os.getenv("TAB_NAME", "April")

SUBJECT_COMPLETED_PHRASE = "Auto Completed Report"
SUBJECT_UPCOMING_PHRASE = "Auto upcoming report"

# Base sheet columns (1-based)
COL_DATE = 2       # B
COL_COMPLETED = 4  # D
COL_SCHEDULED = 5  # E

GRAPH = "https://graph.microsoft.com/v1.0"


# -------------------- Small helpers --------------------
def must_env(name: str) -> str:
    v = os.getenv(name)
    if not v:
        raise RuntimeError(f"Missing environment variable: {name}")
    return v


def parse_dt(dt_str: str) -> datetime:
    if not dt_str:
        return datetime(1970, 1, 1, tzinfo=timezone.utc)
    if dt_str.endswith("Z"):
        dt_str = dt_str.replace("Z", "+00:00")
    return datetime.fromisoformat(dt_str)


def parse_money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[^0-9.\-]", "", str(val))
    return float(s) if s else 0.0


def try_parse_any_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def find_col_idx(header, target_names_lower):
    h = [str(x).strip().lower() for x in header]

    for i, name in enumerate(h):
        if name in target_names_lower:
            return i

    for i, name in enumerate(h):
        for t in target_names_lower:
            if t in name:
                return i

    return None


# -------------------- Auth (App-only) --------------------
def get_token() -> str:
    tenant_id = os.getenv("tenant_id")
    client_id = os.getenv("client_id")
    client_secret = os.getenv("client_secret")

    app = ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )

    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )

    if "access_token" not in result:
        raise RuntimeError(
            f"Token error: {result.get('error')} - {result.get('error_description')}"
        )

    return result["access_token"]


def graph_get(token: str, url: str, params: Optional[dict] = None) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }

    r = requests.get(url, headers=headers, params=params, timeout=60)

    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()

    return r.json()


def graph_get_bytes(token: str, url: str, params: Optional[dict] = None) -> bytes:
    headers = {
        "Authorization": f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }

    r = requests.get(url, headers=headers, params=params, timeout=120)

    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()

    return r.content


def graph_put_bytes(token: str, url: str, content: bytes, content_type: str) -> Dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": content_type,
    }

    r = requests.put(url, headers=headers, data=content, timeout=180)

    if not r.ok:
        try:
            print("Graph error payload:", r.json())
        except Exception:
            print("Graph error text:", r.text)
        r.raise_for_status()

    return r.json()


# -------------------- Outlook: latest email + attachment --------------------
def latest_message_for_subject(token: str, mailbox_upn: str, subject_phrase: str) -> Optional[Dict[str, Any]]:
    """
    Uses $search (contains-style) and sorts locally.
    """
    url = f"{GRAPH}/users/{mailbox_upn}/mailFolders/Inbox/messages"
    params = {
        "$select": "id,subject,receivedDateTime,from,hasAttachments",
        "$top": "25",
        "$search": f"\"{subject_phrase}\"",
    }

    data = graph_get(token, url, params=params)
    msgs: List[Dict[str, Any]] = data.get("value", [])

    phrase = subject_phrase.lower()
    candidates = [m for m in msgs if phrase in (m.get("subject") or "").lower()]

    if not candidates:
        return None

    candidates.sort(key=lambda m: parse_dt(m.get("receivedDateTime", "")), reverse=True)
    return candidates[0]


def get_first_xlsx_attachment_from_message(
    token: str,
    mailbox_upn: str,
    message_id: str
) -> Tuple[Optional[str], Optional[bytes]]:
    """
    Returns (filename, bytes) for the first .xlsx attachment on the message.
    Handles:
    - contentBytes if present
    - fallback to /$value download
    """
    url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments"
    data = graph_get(token, url, params={"$top": "50"})
    atts = data.get("value", [])

    for a in atts:
        name = (a.get("name") or "")
        if name.lower().endswith(".xlsx"):
            cb = a.get("contentBytes")
            if cb:
                return name, base64.b64decode(cb)

            att_id = a.get("id")
            if att_id:
                raw_url = f"{GRAPH}/users/{mailbox_upn}/messages/{message_id}/attachments/{att_id}/$value"
                b = graph_get_bytes(token, raw_url)
                return name, b

    return None, None


# -------------------- SharePoint Excel: download/edit/upload --------------------
def download_sharepoint_excel(token: str, drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    return graph_get_bytes(token, url)


def upload_sharepoint_excel(token: str, drive_id: str, item_id: str, content: bytes) -> None:
    url = f"{GRAPH}/drives/{drive_id}/items/{item_id}/content"
    graph_put_bytes(
        token,
        url,
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -------------------- Attachment parsing --------------------
def read_xlsx_first_sheet_rows(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])

    return rows


def extract_date_from_filename(fname: str) -> date:
    if not fname:
        raise ValueError("Missing filename for date extraction.")

    m = re.search(r"(\d{1,4})[._-](\d{1,2})[._-](\d{1,4})", fname)
    if not m:
        raise ValueError(f"Could not extract date from filename: {fname}")

    a, b, c = m.groups()
    nums = list(map(int, (a, b, c)))

    if nums[0] > 31:
        yyyy, mm, dd = nums[0], nums[1], nums[2]
    else:
        mm, dd, yy = nums
        yyyy = 2000 + yy if yy < 100 else yy

    return date(yyyy, mm, dd)


def subtotal_by_date_from_rows_upcoming(rows):
    if not rows or len(rows) < 2:
        raise ValueError("Attachment sheet is empty or missing header/body.")

    header = rows[0]
    body = rows[1:]

    date_col = find_col_idx(header, {"next appt start date"})
    sub_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if date_col is None or sub_col is None:
        raise ValueError(f"Could not find required columns in Upcoming. Header: {header}")

    totals_by_date = {}
    dates_seen = []

    for r in body:
        if len(r) <= max(date_col, sub_col):
            continue

        d = try_parse_any_date(r[date_col])
        if not d:
            continue

        subtotal = parse_money(r[sub_col])
        totals_by_date[d] = totals_by_date.get(d, 0.0) + subtotal
        dates_seen.append(d)

    if not dates_seen:
        raise ValueError("No valid dates found in Upcoming attachment.")

    today_in_file = min(dates_seen)
    totals_by_date = {k: round(v, 2) for k, v in totals_by_date.items()}

    return today_in_file, totals_by_date


def completed_values_from_rows(rows):
    if not rows or len(rows) < 2:
        raise ValueError("Attachment sheet is empty or missing header/body.")

    header = rows[0]
    body = rows[1:]

    sub_col = find_col_idx(header, {"jobs subtotal", "subtotal"})

    if sub_col is None:
        raise ValueError(f"Could not find subtotal column. Header: {header}")

    total = 0.0

    for r in body:
        if len(r) <= sub_col:
            continue

        v = r[sub_col]
        if v in (None, "", " "):
            continue

        total += parse_money(v)

    return round(total, 2)


# -------------------- SharePoint workbook updates --------------------
def build_sheet_date_row_map_xl(ws, date_col: int) -> Dict[date, int]:
    """
    Reads a date column and returns {date: row_number}
    Assumes row 1 is header.
    """
    mapping = {}
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        v = ws.cell(row=row, column=date_col).value
        d = try_parse_any_date(v)
        if d:
            mapping[d] = row

    return mapping


def apply_upcoming_to_workbook(wb, sheet_name: str, today_in_file: date, totals_by_date: Dict[date, float]) -> int:
    ws = wb[sheet_name]
    base_date_row = build_sheet_date_row_map_xl(ws, COL_DATE)

    updated_cells = 0

    for d, total in totals_by_date.items():
        if d <= today_in_file:
            continue

        row = base_date_row.get(d)
        if not row:
            continue

        ws.cell(row=row, column=COL_SCHEDULED).value = total
        updated_cells += 1

    return updated_cells


def apply_completed_to_workbook(wb, sheet_name: str, file_date: date, global_completed_value: float) -> int:
    ws = wb[sheet_name]
    base_date_row = build_sheet_date_row_map_xl(ws, COL_DATE)
    base_row = base_date_row.get(file_date)

    updates = 0

    if base_row:
        ws.cell(row=base_row, column=COL_COMPLETED).value = global_completed_value
        ws.cell(row=base_row, column=COL_SCHEDULED).value = ""
        updates += 2

    return updates


# -------------------- Main --------------------
def main():
    token = get_token()

    mailbox_upn = "apatil@glassdoctordfw.com"
    drive_id = os.getenv("drive_id")
    file_item_id = os.getenv("file_item_id")

    # 1) Get latest messages
    up_msg = latest_message_for_subject(token, mailbox_upn, SUBJECT_UPCOMING_PHRASE)
    c_msg = latest_message_for_subject(token, mailbox_upn, SUBJECT_COMPLETED_PHRASE)

    if not up_msg and not c_msg:
        print("No matching emails found for Upcoming or Completed.")
        return

    # 2) Download the SharePoint workbook once
    xls_bytes = download_sharepoint_excel(token, drive_id, file_item_id)
    wb = load_workbook(io.BytesIO(xls_bytes))

    if TAB_NAME not in wb.sheetnames:
        raise RuntimeError(f"Tab '{TAB_NAME}' not found in workbook. Found: {wb.sheetnames}")

    # 3) Upcoming: attachment -> parse -> apply
    if up_msg:
        up_id = up_msg["id"]
        up_fname, up_content = get_first_xlsx_attachment_from_message(token, mailbox_upn, up_id)

        if up_content:
            up_rows = read_xlsx_first_sheet_rows(up_content)
            up_today, totals_by_date = subtotal_by_date_from_rows_upcoming(up_rows)
            n = apply_upcoming_to_workbook(wb, TAB_NAME, up_today, totals_by_date)
            print(f"[Upcoming] {up_fname} | file_today={up_today} | updated cells={n}")
        else:
            print("[Upcoming] Email found but no .xlsx attachment.")
    else:
        print("[Upcoming] No matching email found.")

    # 4) Completed: attachment -> parse -> apply
    if c_msg:
        c_id = c_msg["id"]
        c_fname, c_content = get_first_xlsx_attachment_from_message(token, mailbox_upn, c_id)

        if c_content:
            c_rows = read_xlsx_first_sheet_rows(c_content)
            file_date = extract_date_from_filename(c_fname)

            global_sum = completed_values_from_rows(c_rows)
            n = apply_completed_to_workbook(wb, TAB_NAME, file_date, global_sum)

            print(
                f"[Completed] {c_fname} | file_date={file_date} | "
                f"global_completed={global_sum} | updated cells={n}"
            )
        else:
            print("[Completed] Email found but no .xlsx attachment.")
    else:
        print("[Completed] No matching email found.")

    # 5) Save workbook back to bytes and upload
    out = io.BytesIO()
    wb.save(out)
    upload_sharepoint_excel(token, drive_id, file_item_id, out.getvalue())

    print("Done. Uploaded updated Excel back to SharePoint.")


if __name__ == "__main__":
    main()
